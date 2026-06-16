#!/usr/bin/env python3
"""v17.11 — 서울·경기·인천 도매사 명부 (xlsx → JSON)
- 4 시트 (서울, 경기남부, 경기북부, 인천) 통합
- 출력: data/wholesalers.json
"""
import json, re, os, sys
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else '/sessions/exciting-beautiful-carson/mnt/Data/서울경기인천_도매사명부_2025.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else '/sessions/exciting-beautiful-carson/mnt/outputs/data/wholesalers.json'

wb = openpyxl.load_workbook(SRC, data_only=True)

SIDO_MAP = {
    '서울': '서울', '경기도': '경기', '경기': '경기', '인천': '인천', '인천시': '인천',
    '경기도 광명': '경기', '경기도 부천': '경기', '경기도 김포': '경기',
}
def detect_sido(addr, sheet):
    if not addr: return sheet
    a = str(addr).strip()
    if a.startswith('서울'): return '서울'
    if a.startswith('인천'): return '인천'
    if a.startswith('경기'): return '경기'
    # sheet hint
    return {'서울':'서울','경기남부':'경기','경기북부':'경기','인천':'인천'}.get(sheet, sheet)

def detect_gu(addr):
    if not addr: return ''
    a = str(addr).strip()
    # 「서울 강남구」 또는 「강남구 봉은사로」 첫 부분 추출
    m = re.search(r'([가-힣]+[구군시])', a)
    return m.group(1) if m else ''

def clean(v):
    if v is None: return ''
    s = str(v).strip().replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s

def clean_name(name):
    """법인명 정규화 — 「(주)가야주류」 → ㈜ 표기 유지"""
    s = clean(name)
    if not s: return s
    s = s.replace('(주)', '㈜').replace('(유)', '㈜').replace('(명)', '㈜').replace('(자)', '㈜')
    return s

def keyword(name):
    """매칭용 키워드 — ㈜ 등 제거"""
    if not name: return ''
    s = re.sub(r'[㈜㈐()\s]', '', str(name))
    return s.lower()

results = []
sheet_stats = {}

# 서울: row3 헤더, row5부터 데이터 — 법인명 col 2, 대표 col 3, 주소 col 4, 전화 col 5, 팩스 col 6, 우편 col 7
ws = wb['서울']
cnt = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row or not row[0] or not isinstance(row[0], (int, float)): continue
    name = clean_name(row[1] if len(row) > 1 else '')
    if not name: continue
    addr = clean(row[3] if len(row) > 3 else '')
    if addr and not addr.startswith('서울'):
        addr = '서울 ' + addr
    results.append({
        'id': 'w-seoul-' + str(int(row[0])),
        'name': name,
        'ceo': clean(row[2] if len(row) > 2 else ''),
        'addr': addr,
        'sido': '서울',
        'gu': detect_gu(addr),
        'tel': clean(row[4] if len(row) > 4 else ''),
        'fax': clean(row[5] if len(row) > 5 else ''),
        'postal': clean(row[6] if len(row) > 6 else ''),
        'sheet': '서울',
        'kw': keyword(name),
    })
    cnt += 1
sheet_stats['서울'] = cnt

# 경기남부: row3 헤더, row4부터 — 서별 col 1, 사업장 col 2, 법인명 col 3, 대표 col 4, 우편 col 5, 지역 col 6, 전화 col 7, FAX col 8
ws = wb['경기남부']
cnt = 0
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or not row[0] or not isinstance(row[0], (int, float)): continue
    name = clean_name(row[3] if len(row) > 3 else '')
    if not name: continue
    addr = clean(row[2] if len(row) > 2 else '')
    region_code = clean(row[6] if len(row) > 6 else '')
    tel_local = clean(row[7] if len(row) > 7 else '')
    tel = (region_code + tel_local).replace(')', ')-') if region_code else tel_local
    results.append({
        'id': 'w-gyeonggi-s-' + str(int(row[0])),
        'name': name,
        'ceo': clean(row[4] if len(row) > 4 else ''),
        'addr': addr,
        'sido': '경기',
        'gu': detect_gu(addr),
        'tel': tel,
        'fax': clean(row[8] if len(row) > 8 else ''),
        'postal': clean(row[5] if len(row) > 5 else ''),
        'sheet': '경기남부',
        'kw': keyword(name),
    })
    cnt += 1
sheet_stats['경기남부'] = cnt

# 경기북부: row2 헤더, row3부터 — 법인명 col 1, 대표 col 2, 사업장 col 3, 우편 col 4, 전화 col 5, 팩스 col 6
ws = wb['경기북부']
cnt = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not row[0] or not isinstance(row[0], (int, float)): continue
    name = clean_name(row[1] if len(row) > 1 else '')
    if not name: continue
    addr = clean(row[3] if len(row) > 3 else '')
    results.append({
        'id': 'w-gyeonggi-n-' + str(int(row[0])),
        'name': name,
        'ceo': clean(row[2] if len(row) > 2 else ''),
        'addr': addr,
        'sido': '경기',
        'gu': detect_gu(addr),
        'tel': clean(row[5] if len(row) > 5 else ''),
        'fax': clean(row[6] if len(row) > 6 else ''),
        'postal': clean(row[4] if len(row) > 4 else ''),
        'sheet': '경기북부',
        'kw': keyword(name),
    })
    cnt += 1
sheet_stats['경기북부'] = cnt

# 인천: row3 헤더, row5부터 — 회사명 col 1, 대표 col 2, 사업자번호 col 3, 사업장주소 col 4, 우편 col 5, 사업장전화 col 6, FAX col 7
ws = wb['인천']
cnt = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row or not row[0] or not isinstance(row[0], (int, float)): continue
    name = clean_name(row[1] if len(row) > 1 else '')
    if not name: continue
    addr = clean(row[4] if len(row) > 4 else '')
    results.append({
        'id': 'w-incheon-' + str(int(row[0])),
        'name': name,
        'ceo': clean(row[2] if len(row) > 2 else ''),
        'addr': addr,
        'sido': detect_sido(addr, '인천'),
        'gu': detect_gu(addr),
        'tel': clean(row[6] if len(row) > 6 else ''),
        'fax': clean(row[7] if len(row) > 7 else ''),
        'postal': clean(row[5] if len(row) > 5 else ''),
        'sheet': '인천',
        'biz_no': clean(row[3] if len(row) > 3 else ''),
        'kw': keyword(name),
    })
    cnt += 1
sheet_stats['인천'] = cnt

import datetime
out = {
    'updated': datetime.datetime.utcnow().isoformat() + 'Z',
    'source': '서울경기인천_도매사명부_2025.xlsx',
    'total': len(results),
    'by_sheet': sheet_stats,
    'by_sido': {},
    'wholesalers': results,
}
for w in results:
    out['by_sido'][w['sido']] = out['by_sido'].get(w['sido'], 0) + 1

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, separators=(',',':'))
print(f"✅ {OUT}")
print(f"  total: {len(results)}")
print(f"  by_sheet: {sheet_stats}")
print(f"  by_sido: {out['by_sido']}")
print(f"  sample: {results[0]['name']} ({results[0]['sido']} {results[0]['gu']})")
